from openpyxl import load_workbook
import json 
wd=load_workbook(filename="pypy.xlsx")
ws=wd.active

json_map={}
json_map["row"]=[]
for x,y,z in ws.iter_rows(min_row=2, max_col=3, max_row=20):
    json_map["row"].append({
        "x":f"{x.value}",
        "y":f"{y.value}",
        "z":f"{z.value}"
    })

with open("data.txt",'w') as outJson:
    json.dump(json_map,outJson,indent=3)


